#!/usr/bin/env python3
"""Build Acquisition planning.xlsx — Google Sheets-compatible decision workbook."""

from __future__ import annotations

import csv
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent
SRC = ROOT / "Acqusition planning.xlsx"
OUT = ROOT / "Acquisition planning.xlsx"
EXPORTS = ROOT / "exports"

# Styles (xlsx skill conventions)
FONT_INPUT = Font(name="Arial", color="0000FF")
FONT_FORMULA = Font(name="Arial", color="000000")
FONT_HEADER = Font(name="Arial", bold=True)
FONT_LINK = Font(name="Arial", color="008000")
FILL_INPUT = PatternFill("solid", fgColor="E8F4FF")
FILL_HEADER = PatternFill("solid", fgColor="F2F2F2")
FILL_WARN = PatternFill("solid", fgColor="FFFF00")
FILL_PASS = PatternFill("solid", fgColor="C6EFCE")
FILL_FAIL = PatternFill("solid", fgColor="FFC7CE")
FILL_WATCH = PatternFill("solid", fgColor="FFEB9C")

VERTICALS = [
    ("Freight Brokerage", 9, 9, 10, 8, 8, 10, 4, 5),
    ("Construction Estimating", 10, 8, 7, 9, 8, 3, 4, 3),
    ("Permit Expediting", 10, 9, 8, 10, 7, 2, 3, 4),
    ("Industrial Distribution", 8, 8, 10, 8, 8, 5, 4, 5),
    ("Specialty Staffing", 7, 7, 7, 7, 8, 10, 8, 7),
    ("Insurance Inspections", 8, 8, 8, 8, 7, 2, 6, 4),
    ("Valuation Services", 8, 10, 9, 7, 9, 2, 3, 5),
    ("Industry Research / Data", 9, 10, 9, 8, 9, 1, 2, 3),
    ("Vertical SaaS", 8, 9, 9, 10, 10, 2, 2, 4),
    ("MSP", 6, 5, 6, 6, 6, 1, 8, 8),
    ("HVAC", 4, 2, 3, 2, 3, 1, 9, 5),
    ("Roofing", 4, 2, 4, 2, 2, 1, 10, 6),
    ("Self Storage", 2, 2, 5, 3, 1, 1, 2, 1),
    ("Laundromat", 1, 1, 3, 2, 1, 1, 2, 1),
]

CATEGORIES_THEMES = [
    (1, "Vertical logistics / dispatch service with software layer",
     "Direct match to quote-to-dispatch, routing, crew assignment, and pricing work.",
     "Labor, claims, seasonality, customer concentration.", ""),
    (2, "Broker / intermediary business with ugly workflow",
     "Fits strength in reducing ambiguity, speeding quote generation, and standardizing handoffs.",
     "You may inherit relationship-heavy revenue that resists systemization.", ""),
    (3, "B2B services business with recurring ops chaos",
     "Good if there is a repetitive intake, scheduling, billing, compliance, or QA loop.",
     "Margin leakage can be hidden until after close.", ""),
    (4, "Niche vertical SaaS with service wrapper",
     "Best if the software is underbuilt and the real product is workflow control.",
     "Pure SaaS may be too slow if it lacks operational leverage.", ""),
    (5, "HR / workforce / compliance admin business",
     "Strong fit because you understand transaction-ready data, integration, and diligence workflows.",
     "Sales cycles and buyer conservatism can slow change.", ""),
]

CATEGORIES_TYPES = [
    (1, "Last-mile / dispatch-heavy logistics, 3PL, freight brokerage, specialized routing",
     "Direct match to quote-to-dispatch, pricing, status, crew/workforce orchestration, and operational AI mindset.",
     "$700k–$2.5M", ""),
    (2, "Field-service or installation business with scheduling complexity",
     "Same bottleneck pattern as logistics: intake, dispatch, exceptions, capacity, and customer status.",
     "$500k–$1.8M", ""),
    (3, "Business brokerage / advisory / transaction services with process drag",
     "Strong fit if the real problem is Excel, email, and manual reporting.",
     "$400k–$1.5M", ""),
    (4, "HR, payroll, diligence, or workforce-integration services",
     "Fits transaction-readiness and integration-first thinking.",
     "$500k–$2.0M", ""),
    (5, "Compliance / trust-sensitive monitoring business",
     "Good if signal quality matters more than feature breadth.",
     "$600k–$2.0M", ""),
    (6, "Generic local service business without workflow leverage",
     "You can own it, but it is mostly management labor, not systems leverage.",
     "Any", ""),
    (7, "Passive-style: parking lots, freight-yard, storage-yard, junk-yard",
     "Cash-flow asset, not product-led operating leverage.",
     "Varies", "Reject unless goal changes to cash-flow investing"),
]


def style_header_row(ws, row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        c = ws.cell(row=row, column=col)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = Alignment(wrap_text=True, vertical="top")


def style_input_cell(cell) -> None:
    cell.font = FONT_INPUT
    cell.fill = FILL_INPUT


def style_formula_cell(cell) -> None:
    cell.font = FONT_FORMULA


def set_col_widths(ws, widths: dict[int, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def add_yes_no_validation(ws, cell_range: str) -> None:
    dv = DataValidation(type="list", formula1='"YES,NO,Unknown"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(cell_range)


def add_score_validation(ws, cell_range: str) -> None:
    dv = DataValidation(type="whole", operator="between", formula1=1, formula2=5, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(cell_range)


def add_status_validation(ws, cell_range: str) -> None:
    dv = DataValidation(
        type="list",
        formula1='"New,Screening,Diligence,LOI,Closed-Won,Passed,Killed"',
        allow_blank=True,
    )
    ws.add_data_validation(dv)
    dv.add(cell_range)


def freeze_header(ws) -> None:
    ws.freeze_panes = "A2"


def sheet_readme(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("00_Readme", 0)
    lines = [
        ("ACQUISITION DECISION SYSTEM", True),
        ("", False),
        ("Current call: NOT YET — bounded secondary experiment only (max 5 hrs/week).", False),
        ("", False),
        ("TAB GUIDE", True),
        ("01_Archetype_Scorecard — vertical fit with formula-driven weighted scores", False),
        ("02_Target_Size_Bands — SDE/EBITDA/EV bands and affordability calculator", False),
        ("03_Operator_Readiness — financing, runway, willingness gates", False),
        ("04_Deal_Pipeline — real listings only", False),
        ("05_Deal_Scorecard — per-deal weighted score + Pursue/Watch/Kill", False),
        ("06_Business_Risk_Diligence — personnel, ops, customer, revenue quality", False),
        ("07_Diligence_Memo — one-page memo template per deal", False),
        ("08_Time_Budget — weekly hours across Path A, Path B, acquisition", False),
        ("09_Decision_Log — dated pass/kill decisions", False),
        ("", False),
        ("THREE GATES (all required)", True),
        ("1. Business you could improve", False),
        ("2. Business you could realistically buy", False),
        ("3. Business you would operate for 3 years", False),
        ("", False),
        ("DEAL SCORECARD WEIGHTS", True),
        ("Operator fit — 25%", False),
        ("AI/product leverage — 20%", False),
        ("Financial feasibility — 25%", False),
        ("Owner dependence / transition — 15%", False),
        ("Operating enjoyment — 10%", False),
        ("Strategic conflict risk — 5%", False),
        ("", False),
        ("LABELS: Pursue (score >= 3.5 and gates pass) | Watch (2.5-3.4 or missing data) | Kill (< 2.5 or hard fail)", False),
        ("", False),
        ("FORMULA APPENDIX (Google Sheets-safe)", True),
        ("Weighted total: =SUMPRODUCT(weights, scores) / SUM(weights)", False),
        ("Decision label: =IF(AND(score>=3.5, gates=\"YES\"), \"Pursue\", IF(OR(score<2.5, hardFail=\"YES\"), \"Kill\", \"Watch\"))", False),
        ("DSCR: =IF(debtService=0, \"\", normalizedEBITDA/debtService)", False),
        ("EV/SDE: =IF(SDE=0, \"\", askingPrice/SDE)", False),
        ("", False),
        ("KILL RULES — pause search if:", True),
        ("• Acquisition > 8 hrs/wk for 2 weeks while Path B outbound < 5 hrs/wk", False),
        ("• New domain listing outside ops-AI/logistics (month-5 relapse)", False),
        ("• LOI without financing pre-approval", False),
        ("• Deal fails all three gates", False),
        ("", False),
        ("See ACQUISITION/README.md and REPENTANCE/acquisition-decision/ for full rules.", False),
    ]
    for i, (text, bold) in enumerate(lines, 1):
        c = ws.cell(row=i, column=1, value=text)
        if bold:
            c.font = Font(name="Arial", bold=True, size=12 if i == 1 else 11)
        else:
            c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 100


def sheet_archetype_scorecard(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("01_Archetype_Scorecard")
    headers = [
        "Vertical", "Workflow", "Info Arbitrage", "Pricing", "Automation",
        "Discovery", "Marketplace", "Labor Burden (-)", "Sales Hero (-)",
        "Weighted Score", "Tier", "Notes",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    # Factor weights row (hidden reference area row 2)
    weights = [0.15, 0.15, 0.15, 0.15, 0.10, 0.10, 0.10, 0.10]
    ws.cell(row=2, column=1, value="Factor weights (ref)")
    for i, w in enumerate(weights, 2):
        c = ws.cell(row=2, column=i, value=w)
        style_formula_cell(c)
    ws.row_dimensions[2].hidden = True

    start = 3
    for idx, row_data in enumerate(VERTICALS):
        r = start + idx
        name = row_data[0]
        scores = row_data[1:]
        ws.cell(row=r, column=1, value=name)
        for j, s in enumerate(scores, 2):
            c = ws.cell(row=r, column=j, value=s)
            style_input_cell(c)
        # Weighted score: positive factors B-G, inverted negatives H-I (10 - value)
        ws.cell(
            row=r, column=10,
            value=(
                f"=ROUND(("
                f"B{r}*$B$2+C{r}*$C$2+D{r}*$D$2+E{r}*$E$2+F{r}*$F$2+G{r}*$G$2+"
                f"(10-H{r})*$H$2+(10-I{r})*$I$2"
                f")/SUM($B$2:$I$2),1)"
            ),
        )
        style_formula_cell(ws.cell(row=r, column=10))
        ws.cell(
            row=r, column=11,
            value=f'=IF(J{r}>=7.5,"Tier 1",IF(J{r}>=5.5,"Tier 2","Tier 3"))',
        )
        style_formula_cell(ws.cell(row=r, column=11))

    # Preserved category themes (below vertical table)
    base = start + len(VERTICALS) + 2
    ws.cell(row=base, column=1, value="CATEGORY THEMES (preserved from original workbook)")
    ws.cell(row=base, column=1).font = FONT_HEADER
    hdr = base + 1
    for col, h in enumerate(["#", "Business type & Theme", "Why it fits", "Main risk", "Target EV"], 1):
        ws.cell(row=hdr, column=col, value=h)
    style_header_row(ws, hdr, 5)
    for i, row in enumerate(CATEGORIES_THEMES):
        r = hdr + 1 + i
        for j, val in enumerate(row, 1):
            ws.cell(row=r, column=j, value=val)

    base2 = hdr + 1 + len(CATEGORIES_THEMES) + 2
    ws.cell(row=base2, column=1, value="ACQUISITION TYPES (preserved)")
    ws.cell(row=base2, column=1).font = FONT_HEADER
    hdr2 = base2 + 1
    for col, h in enumerate(["#", "Acquisition type", "Why it fits", "Target EBITDA", "Target EV / Note"], 1):
        ws.cell(row=hdr2, column=col, value=h)
    style_header_row(ws, hdr2, 5)
    for i, row in enumerate(CATEGORIES_TYPES):
        r = hdr2 + 1 + i
        for j, val in enumerate(row, 1):
            ws.cell(row=r, column=j, value=val)

    set_col_widths(ws, {1: 28, 2: 12, 3: 12, 4: 12, 5: 12, 6: 12, 7: 12, 8: 14, 9: 14, 10: 14, 11: 10, 12: 20})
    freeze_header(ws)


def sheet_target_size_bands(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("02_Target_Size_Bands")
    rows = [
        ("TARGET SIZE BANDS — small individual acquirer", ""),
        ("Band", "SDE / EBITDA", "EV", "Notes"),
        ("Preferred", "$150K–$500K", "$500K–$2.0M", "Default search range"),
        ("Stretch", "$500K–$750K", "$2.0M–$3.0M", "Requires lender pre-qual + seller financing"),
        ("Avoid", "< $150K or > $750K EBITDA", "> $3.0M EV", "Too small (job) or too large without capital"),
        ("", "", "", ""),
        ("AFFORDABILITY CALCULATOR (fill yellow cells)", ""),
        ("Field", "Value", "Formula / Threshold", "Pass?"),
        ("Asking price / EV ($)", None, "INPUT", None),
        ("Seller-reported SDE ($)", None, "INPUT", None),
        ("Seller-reported EBITDA ($)", None, "INPUT", None),
        ("Normalized buyer SDE ($)", None, "INPUT — use for multiples", None),
        ("Normalized buyer EBITDA ($)", None, "INPUT — use for debt service", None),
        ("Down payment %", 0.20, "INPUT (default 20%)", None),
        ("Seller financing %", 0.10, "INPUT", None),
        ("Interest rate (annual)", 0.08, "INPUT", None),
        ("Loan term (years)", 10, "INPUT", None),
        ("Liquid capital available ($)", None, "INPUT — fill in private copy only", None),
        ("Minimum personal income needed ($/yr)", None, "INPUT", None),
        ("", "", "", ""),
        ("EV / SDE (x)", None, "=IF(B11=0,\"\",B9/B11)", None),
        ("EV / EBITDA (x)", None, "=IF(B13=0,\"\",B9/B13)", None),
        ("Cash required ($)", None, "=B9*B15", None),
        ("Loan principal ($)", None, "=B9*(1-B15-B16)", None),
        ("Annual debt service ($)", None, "=IF(AND(B18>0,B19>0),PMT(B17,B19,-B22),\"\")", None),
        ("DSCR (x)", None, "=IF(B24=0,\"\",B13/B24)", None),
        ("Post-debt owner cash flow ($)", None, "=IF(B13=\"\",\"\",B13-B24)", None),
        ("Size band", None, "formula", None),
        ("Financing hard fail?", None, "formula", None),
    ]
    for i, row in enumerate(rows, 1):
        for j, val in enumerate(row, 1):
            c = ws.cell(row=i, column=j, value=val)
            if i in (1, 7):
                c.font = FONT_HEADER
            if i == 2 or i == 8:
                style_header_row(ws, i, 4)

    # Input rows 9-20 (B column)
    input_rows = [9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20]
    for r in input_rows:
        style_input_cell(ws.cell(row=r, column=2))
        if r in (15, 16, 17, 18, 19):
            ws.cell(row=r, column=2).fill = FILL_WARN

    # Formulas — B9 asking, B12 norm SDE, B13 norm EBITDA, B14 down%, B15 seller%, B16 rate, B17 term
    ws["B21"] = "=IF(B12=0,\"\",B9/B12)"
    ws["B22"] = "=IF(B13=0,\"\",B9/B13)"
    ws["B23"] = "=B9*B14"
    ws["B24"] = "=B9*(1-B14-B15)"
    ws["B25"] = "=IF(AND(B16>0,B17>0),ABS(PMT(B16,B17,-B24)),\"\")"
    ws["B26"] = "=IF(B25=0,\"\",B13/B25)"
    ws["B27"] = "=IF(B13=\"\",\"\",B13-B25)"
    ws["B28"] = (
        '=IF(OR(B13<150000,B13>750000,B9>3000000),"Avoid",'
        'IF(AND(B13>=150000,B13<=500000,B9<=2000000),"Preferred",'
        'IF(AND(B13>500000,B13<=750000,B9<=3000000),"Stretch","Review")))'
    )
    ws["B29"] = (
        '=IF(OR(B18=\"\",B23>B18,B26<1.25,B27<B19),"YES","NO")'
    )
    for ref in ["B21", "B22", "B23", "B24", "B25", "B26", "B27", "B28", "B29"]:
        style_formula_cell(ws[ref])

    set_col_widths(ws, {1: 32, 2: 18, 3: 28, 4: 12})
    freeze_header(ws)


def sheet_operator_readiness(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("03_Operator_Readiness")
    ws.cell(row=1, column=1, value="OPERATOR READINESS — default: NOT YET")
    ws.cell(row=1, column=1).font = FONT_HEADER

    sections = [
        ("1. Capital and financing", ["Item", "Your answer", "Pass?"], [
            ("Liquid equity for down payment ($)", "", ""),
            ("Additional investable without ruin ($)", "", ""),
            ("SBA 7(a) or seller note pre-qualified?", "Unknown", ""),
            ("Max monthly debt service affordable ($)", "", ""),
            ("Spouse/partner alignment on personal guarantee", "Unknown", ""),
        ]),
        ("2. Runway and income pressure", ["Item", "Your answer", "Pass?"], [
            ("Months of personal runway at current burn", "", ""),
            ("Minimum monthly income needed ($)", "", ""),
            ("Can you go 6-12 months with $0 salary during search?", "NO", ""),
            ("Fractional income still required ($/mo)", "", ""),
        ]),
        ("3. Operator willingness (1-5)", ["Activity", "Score (1-5)", "Pass?"], [
            ("Reviewing P&L weekly", "", ""),
            ("Managing payroll / HR issues", "", ""),
            ("Vendor negotiations", "", ""),
            ("Seller transition meetings", "", ""),
            ("Firing underperformers", "", ""),
            ("Same customer problem 200 times", "", ""),
            ("Product/roadmap decisions post-close", "", ""),
        ]),
        ("4. Seller transition", ["Item", "Plan", "Pass?"], [
            ("Weeks seller stays post-close", "", ""),
            ("Non-compete geography", "", ""),
            ("Customer intro commitments", "", ""),
            ("Your hours/week during transition", "", ""),
        ]),
        ("5. Time budget", ["Activity", "Hrs/week budget", "Actual (4-wk avg)"], [
            ("Path B: founder outreach / apps", 12, ""),
            ("Path A: Quotely distribution", 7, ""),
            ("Acquisition: search + diligence", 5, ""),
        ]),
        ("6. Professional support", ["Resource", "Secured?", "Pass?"], [
            ("SMB-focused attorney (LOI, APA)", "NO", ""),
            ("QoE accountant or diligence help", "NO", ""),
            ("Business broker relationship", "NO", ""),
            ("SBA lender contact", "NO", ""),
        ]),
    ]

    row = 3
    willingness_start = willingness_end = 0
    time_path_b_row = time_acq_row = 0
    for title, hdr, items in sections:
        ws.cell(row=row, column=1, value=title).font = FONT_HEADER
        row += 1
        for j, h in enumerate(hdr, 1):
            ws.cell(row=row, column=j, value=h)
        style_header_row(ws, row, 3)
        row += 1
        section_start = row
        for item in items:
            ws.cell(row=row, column=1, value=item[0])
            c_ans = ws.cell(row=row, column=2, value=item[1] if item[1] else None)
            style_input_cell(c_ans)
            if title.startswith("3."):
                add_score_validation(ws, f"B{row}")
            if "pre-qualified" in item[0] or "guarantee" in item[0] or "Secured" in hdr[1]:
                add_yes_no_validation(ws, f"B{row}")
            row += 1
        section_end = row - 1
        if title.startswith("3."):
            willingness_start, willingness_end = section_start, section_end
        if title.startswith("5."):
            time_path_b_row = section_start
            time_acq_row = section_start + 2
        # Pass formula for section (col C)
        for r in range(section_start, section_end + 1):
            if title.startswith("3."):
                ws.cell(row=r, column=3, value=f'=IF(B{r}="","",IF(B{r}>=3.5,"Pass","Fail"))')
            elif title.startswith("5."):
                pass  # no pass col for time budget
            else:
                ws.cell(row=r, column=3, value=f'=IF(B{r}="","Incomplete",IF(OR(B{r}="NO",B{r}="Unknown"),"Fail","Pass"))')
            style_formula_cell(ws.cell(row=r, column=3))
        row += 1

    # Summary block
    row += 1
    ws.cell(row=row, column=1, value="READINESS SUMMARY").font = FONT_HEADER
    row += 1
    fin_row = row
    ws.cell(row=fin_row, column=1, value="Financing documented?")
    ws.cell(row=fin_row, column=2, value='=IF(COUNTIF(C5:C9,"Fail")>0,"Fail",IF(COUNTIF(C5:C9,"Incomplete")>0,"Incomplete","Pass"))')
    run_row = fin_row + 1
    ws.cell(row=run_row, column=1, value="Runway documented?")
    ws.cell(row=run_row, column=2, value='=IF(COUNTIF(C12:C15,"Fail")>0,"Fail",IF(COUNTIF(C12:C15,"Incomplete")>0,"Incomplete","Pass"))')
    will_avg_row = run_row + 1
    ws.cell(row=will_avg_row, column=1, value="Operator willingness avg")
    ws.cell(row=will_avg_row, column=2, value=f'=IF(COUNT(B{willingness_start}:B{willingness_end})=0,"",ROUND(AVERAGE(B{willingness_start}:B{willingness_end}),1))')
    will_pass_row = will_avg_row + 1
    ws.cell(row=will_pass_row, column=1, value="Willingness pass?")
    ws.cell(row=will_pass_row, column=2, value=f'=IF(B{will_avg_row}="","",IF(B{will_avg_row}>=3.5,"Pass","Fail"))')
    kill_row = will_pass_row + 1
    ws.cell(row=kill_row, column=1, value="Time budget kill flag?")
    ws.cell(
        row=kill_row, column=2,
        value=f'=IF(AND(C{time_acq_row}>8,C{time_path_b_row}<5),"YES — pause acquisition","NO")',
    )
    overall_row = kill_row + 1
    ws.cell(row=overall_row, column=1, value="OVERALL").font = FONT_HEADER
    ws.cell(row=overall_row, column=2, value=f'=IF(OR(B{fin_row}="Fail",B{run_row}="Fail",B{will_pass_row}="Fail",B{kill_row}="YES — pause acquisition"),"NOT YET","Review — may proceed to active search")')
    for r in range(fin_row, overall_row + 1):
        style_formula_cell(ws.cell(row=r, column=2))

    set_col_widths(ws, {1: 42, 2: 22, 3: 18, 4: 16})
    freeze_header(ws)


def sheet_deal_pipeline(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("04_Deal_Pipeline")
    headers = [
        "Deal ID", "Deal name", "Source", "Archetype", "Vertical",
        "Asking price ($)", "Seller SDE ($)", "Seller EBITDA ($)",
        "Normalized SDE ($)", "Normalized EBITDA ($)",
        "Seller financing %", "Status", "Notes", "Date added",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    # 5 blank template rows
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f"D{r-1:03d}")
        for col in range(2, 15):
            style_input_cell(ws.cell(row=r, column=col))
        add_status_validation(ws, f"L{r}")

    set_col_widths(ws, {1: 8, 2: 24, 3: 14, 4: 14, 5: 20, 6: 14, 7: 12, 8: 14, 9: 14, 10: 16, 11: 14, 12: 12, 13: 30, 14: 12})
    freeze_header(ws)


def sheet_deal_scorecard(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("05_Deal_Scorecard")
    headers = [
        "Deal ID", "Deal name", "Operator fit (1-5)", "AI/product leverage (1-5)",
        "Financial feasibility (1-5)", "Owner dependence (1-5)", "Operating enjoyment (1-5)",
        "Strategic conflict (1-5)", "Weighted score", "Gate: improve?", "Gate: buy?",
        "Gate: operate 3yr?", "Financing documented?", "Path B displaced?",
        "Risk red flags (#)", "All gates pass?", "Hard fail?", "Decision",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    # Weights reference row 2 (hidden)
    weights = [0.25, 0.20, 0.25, 0.15, 0.10, 0.05]
    labels = ["", "Weights", "Op fit", "AI lev", "Fin feas", "Owner dep", "Enjoy", "Strat"]
    for j, (lbl, w) in enumerate(zip(labels, [""] + weights), 1):
        ws.cell(row=2, column=j, value=lbl if j == 1 else (w if j > 2 else labels[j - 1]))
    ws.row_dimensions[2].hidden = True
    ws.cell(row=2, column=3, value=0.25)
    ws.cell(row=2, column=4, value=0.20)
    ws.cell(row=2, column=5, value=0.25)
    ws.cell(row=2, column=6, value=0.15)
    ws.cell(row=2, column=7, value=0.10)
    ws.cell(row=2, column=8, value=0.05)

    for r in range(3, 8):
        ws.cell(row=r, column=1, value=f"D{r-2:03d}")
        for col in range(2, 15):
            style_input_cell(ws.cell(row=r, column=col))
        add_score_validation(ws, f"C{r}:H{r}")
        add_yes_no_validation(ws, f"J{r}:N{r}")
        # Weighted score
        ws.cell(
            row=r, column=9,
            value=f"=IF(COUNT(C{r}:H{r})<6,\"\",ROUND(SUMPRODUCT(C{r}:H{r},$C$2:$H$2)/SUM($C$2:$H$2),2))",
        )
        style_formula_cell(ws.cell(row=r, column=9))
        # Risk red flags — link to diligence sheet row (same deal row)
        dr = r  # diligence row aligns
        ws.cell(
            row=r, column=15,
            value=f"='06_Business_Risk_Diligence'!V{dr}",
        )
        style_formula_cell(ws.cell(row=r, column=15))
        # All gates pass
        ws.cell(row=r, column=16, value=f'=IF(COUNTIF(J{r}:N{r},"NO")>0,"NO","YES")')
        style_formula_cell(ws.cell(row=r, column=16))
        # Hard fail
        ws.cell(
            row=r, column=17,
            value=f'=IF(OR(N{r}="YES",O{r}>=3,M{r}="NO"),"YES","NO")',
        )
        style_formula_cell(ws.cell(row=r, column=17))
        # Decision
        ws.cell(
            row=r, column=18,
            value=(
                f'=IF(AND(I{r}>=3.5,P{r}="YES",Q{r}="NO"),"Pursue",'
                f'IF(OR(I{r}<2.5,Q{r}="YES"),"Kill","Watch"))'
            ),
        )
        style_formula_cell(ws.cell(row=r, column=18))

    set_col_widths(ws, {1: 8, 2: 22, 9: 12, 15: 12, 16: 12, 17: 10, 18: 10})
    freeze_header(ws)


def sheet_business_risk_diligence(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("06_Business_Risk_Diligence")
    flag_headers = [
        "Deal ID", "Deal name",
        "Owner is primary seller/dispatcher?", "No second-in-command?",
        "Key employee departure risk?", "High turnover / heroics?",
        "Underpaid relatives or misclassified contractors?",
        "SOP / tribal knowledge risk?", "Labor burden / OT exposure?",
        "Cannot explain gross margin by job type?", "Poor cash conversion?",
        "Systems messy and data absent?", "Deferred maintenance / claims?",
        "Top customer > 20% revenue?", "Top 3 customers > 40%?",
        "Founder owns all relationships?", "Revenue re-sold each cycle?",
        "Paid ads with unclear CAC?", "Ops trust problems in reviews?",
        "Add-backs low confidence?", "Working capital surprise?",
        "Transition < 60 days seller support?", "Red flag count", "Risk band",
    ]
    for col, h in enumerate(flag_headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(flag_headers))

    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f"D{r-1:03d}")
        ws.cell(row=r, column=2, value="")
        style_input_cell(ws.cell(row=r, column=2))
        for col in range(3, 22):
            c = ws.cell(row=r, column=col, value="")
            style_input_cell(c)
            add_yes_no_validation(ws, f"{get_column_letter(col)}{r}")
        ws.cell(
            row=r, column=22,
            value=f'=COUNTIF(C{r}:U{r},"YES")',
        )
        style_formula_cell(ws.cell(row=r, column=22))
        ws.cell(
            row=r, column=23,
            value=f'=IF(V{r}>=5,"High",IF(V{r}>=2,"Medium","Low"))',
        )
        style_formula_cell(ws.cell(row=r, column=23))

    set_col_widths(ws, {1: 8, 2: 20, 22: 12, 23: 10})
    freeze_header(ws)


def sheet_diligence_memo(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("07_Diligence_Memo")
    fields = [
        "Deal ID", "Deal name", "Date", "Archetype", "Asking / SDE / EBITDA",
        "Why this could work (your wedge)", "Why it might fail",
        "Normalized buyer economics", "Transition plan (60+ days?)",
        "Compare to one week of Path B outbound", "Decision", "Next action",
    ]
    for col, h in enumerate(fields, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(fields))
    for r in range(2, 7):
        for col in range(1, 13):
            style_input_cell(ws.cell(row=r, column=col))
    set_col_widths(ws, {1: 8, 2: 20, 5: 24, 6: 36, 7: 36, 8: 24, 11: 12, 12: 24})
    freeze_header(ws)


def sheet_time_budget(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("08_Time_Budget")
    headers = ["Week ending", "Path B outbound (hrs)", "Path A Quotely (hrs)", "Acquisition (hrs)", "Total", "Kill flag?"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))
    for r in range(2, 14):
        for col in range(1, 4):
            style_input_cell(ws.cell(row=r, column=col))
        ws.cell(row=r, column=5, value=f"=SUM(B{r}:D{r})")
        style_formula_cell(ws.cell(row=r, column=5))
        ws.cell(row=r, column=6, value=f'=IF(AND(D{r}>8,B{r}<5),"YES — pause acquisition","NO")')
        style_formula_cell(ws.cell(row=r, column=6))
    set_col_widths(ws, {1: 14, 2: 18, 3: 18, 4: 16, 5: 10, 6: 22})
    freeze_header(ws)


def sheet_decision_log(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("09_Decision_Log")
    headers = ["Date", "Deal ID", "Deal name", "Decision", "Weighted score", "Reason", "Next action", "Revisit date"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))
    for r in range(2, 22):
        for col in range(1, 9):
            style_input_cell(ws.cell(row=r, column=col))
    dv = DataValidation(type="list", formula1='"Pursue,Watch,Pass,Kill,Parked"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("D2:D21")
    set_col_widths(ws, {1: 12, 2: 8, 3: 22, 4: 10, 5: 12, 6: 40, 7: 24, 8: 12})
    freeze_header(ws)


def export_csvs(wb: openpyxl.Workbook) -> None:
    EXPORTS.mkdir(exist_ok=True)

    def write_sheet_csv(sheet_name: str, path: Path) -> None:
        ws = wb[sheet_name]
        with path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
                if any(v is not None and v != "" for v in row):
                    writer.writerow(row)

    write_sheet_csv("04_Deal_Pipeline", EXPORTS / "deal_pipeline.csv")
    write_sheet_csv("09_Decision_Log", EXPORTS / "decision_log.csv")


def verify_workbook(path: Path) -> list[str]:
    issues = []
    wb = openpyxl.load_workbook(path, data_only=False)
    bad_tokens = ("#REF!", "#NAME?", "#DIV/0!", "#VALUE!", "#N/A")
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and any(t in v for t in bad_tokens):
                    issues.append(f"{name}!{cell.coordinate}: {v}")
    return issues


def main() -> None:
    wb = openpyxl.Workbook()
    # Remove default sheet after creating ours
    default = wb.active
    sheet_readme(wb)
    sheet_archetype_scorecard(wb)
    sheet_target_size_bands(wb)
    sheet_operator_readiness(wb)
    sheet_deal_pipeline(wb)
    sheet_deal_scorecard(wb)
    sheet_business_risk_diligence(wb)
    sheet_diligence_memo(wb)
    sheet_time_budget(wb)
    sheet_decision_log(wb)
    if default.title == "Sheet" and default.max_row == 1 and default.max_column == 1:
        wb.remove(default)

    wb.save(OUT)
    export_csvs(wb)
    issues = verify_workbook(OUT)
    if issues:
        print("FORMULA ISSUES:")
        for i in issues:
            print(" ", i)
    else:
        print(f"Saved {OUT} ({len(wb.sheetnames)} sheets)")
        print(f"Exported CSVs to {EXPORTS}")


if __name__ == "__main__":
    main()
